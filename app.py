import os
import io
import json
from datetime import datetime
from flask import Flask, request, jsonify, render_template, send_file
from flask_cors import CORS
from dotenv import load_dotenv
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

load_dotenv()

app = Flask(__name__)
CORS(app)
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024

SCOPES = ['https://www.googleapis.com/auth/drive']
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp', 'bmp', 'tiff', 'svg'}
FOLDER_NAME = 'Amazon-Product-Images'

MIME_MAP = {
    'jpg': 'image/jpeg', 'jpeg': 'image/jpeg',
    'png': 'image/png', 'gif': 'image/gif',
    'webp': 'image/webp', 'bmp': 'image/bmp',
    'tiff': 'image/tiff', 'svg': 'image/svg+xml'
}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_drive_service():
    secret_file = '/etc/secrets/google_credentials.json'
    creds_json = os.getenv('GOOGLE_CREDENTIALS_JSON')
    if creds_json:
        creds_dict = json.loads(creds_json)
        creds = service_account.Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
    elif os.path.exists(secret_file):
        creds = service_account.Credentials.from_service_account_file(secret_file, scopes=SCOPES)
    else:
        creds = service_account.Credentials.from_service_account_file('credentials.json', scopes=SCOPES)
    return build('drive', 'v3', credentials=creds, cache_discovery=False)

def get_or_create_folder(service):
    folder_id = os.getenv('GOOGLE_DRIVE_FOLDER_ID')
    if folder_id:
        return folder_id
    query = f"name='{FOLDER_NAME}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
    res = service.files().list(q=query, fields='files(id)').execute()
    files = res.get('files', [])
    if files:
        return files[0]['id']
    meta = {'name': FOLDER_NAME, 'mimeType': 'application/vnd.google-apps.folder'}
    folder = service.files().create(body=meta, fields='id').execute()
    fid = folder['id']
    service.permissions().create(fileId=fid, body={'type': 'anyone', 'role': 'reader'}, fields='id').execute()
    return fid

def make_public(service, file_id):
    service.permissions().create(fileId=file_id, body={'type': 'anyone', 'role': 'reader'}, fields='id').execute()

def direct_url(file_id):
    return f"https://lh3.googleusercontent.com/d/{file_id}"

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_images():
    if 'images' not in request.files:
        return jsonify({'error': 'No images provided'}), 400
    files = request.files.getlist('images')
    if not files or all(f.filename == '' for f in files):
        return jsonify({'error': 'No files selected'}), 400
    try:
        service = get_drive_service()
        folder_id = get_or_create_folder(service)
    except Exception as e:
        return jsonify({'error': f'Google Drive connection failed: {str(e)}'}), 500
    results, errors = [], []
    for file in files:
        if file.filename == '':
            continue
        if not allowed_file(file.filename):
            errors.append({'filename': file.filename, 'error': 'File type not supported'})
            continue
        try:
            ext = file.filename.rsplit('.', 1)[1].lower()
            mime = MIME_MAP.get(ext, 'image/jpeg')
            data = file.read()
            meta = {'name': file.filename, 'parents': [folder_id]}
            media = MediaIoBaseUpload(io.BytesIO(data), mimetype=mime, resumable=False)
            uploaded = service.files().create(body=meta, media_body=media, fields='id, name, size').execute()
            fid = uploaded['id']
            make_public(service, fid)
            results.append({
                'filename': file.filename,
                'url': direct_url(fid),
                'file_id': fid,
                'size': int(uploaded.get('size', 0)),
                'format': ext.upper(),
            })
        except Exception as e:
            errors.append({'filename': file.filename, 'error': str(e)})
    return jsonify({'success': results, 'errors': errors})

@app.route('/export-excel', methods=['POST'])
def export_excel():
    data = request.json
    if not data or not data.get('images'):
        return jsonify({'error': 'No data provided'}), 400
    images = data['images']
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Amazon Image URLs'
    ORANGE, YELLOW, NAVY, WHITE, GREY = 'FF6B2B', 'FEBD69', '0F1923', 'FFFFFF', 'F8F9FB'
    thin = Side(style='thin', color='E0E0E0')
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.merge_cells('A1:D1')
    t = ws['A1']
    t.value = f'Amazon Product Image URLs  —  {datetime.now().strftime("%B %d, %Y  %H:%M")}'
    t.font = Font(name='Calibri', bold=True, size=13, color=WHITE)
    t.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
    t.alignment = Alignment(horizontal='left', vertical='center', indent=2)
    ws.row_dimensions[1].height = 32
    ws.merge_cells('A2:D2')
    s = ws['A2']
    s.value = f'{len(images)} image(s)  —  Ready for Amazon Seller Central bulk import'
    s.font = Font(name='Calibri', size=10, italic=True, color='555555')
    s.fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type='solid')
    s.alignment = Alignment(horizontal='left', vertical='center', indent=2)
    ws.row_dimensions[2].height = 22
    for i, (h, w, l) in enumerate(zip(['#', 'Image Name', 'Image URL (Direct)', 'Format'], [5, 38, 90, 10], ['A', 'B', 'C', 'D']), 1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = Font(name='Calibri', bold=True, size=11, color=WHITE)
        c.fill = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type='solid')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = bd
        ws.column_dimensions[l].width = w
    ws.row_dimensions[3].height = 26
    for idx, img in enumerate(images, 1):
        r = idx + 3
        fc = GREY if idx % 2 == 0 else WHITE
        rf = PatternFill(start_color=fc, end_color=fc, fill_type='solid')
        c1 = ws.cell(row=r, column=1, value=idx)
        c1.font = Font(name='Calibri', size=10, color='888888')
        c1.fill = rf; c1.alignment = Alignment(horizontal='center', vertical='center'); c1.border = bd
        c2 = ws.cell(row=r, column=2, value=img.get('filename', ''))
        c2.font = Font(name='Calibri', size=10, bold=True)
        c2.fill = rf; c2.alignment = Alignment(horizontal='left', vertical='center', indent=1); c2.border = bd
        url = img.get('url', '')
        c3 = ws.cell(row=r, column=3, value=url)
        c3.hyperlink = url
        c3.font = Font(name='Calibri', size=10, color='0563C1', underline='single')
        c3.fill = rf; c3.alignment = Alignment(horizontal='left', vertical='center', indent=1); c3.border = bd
        c4 = ws.cell(row=r, column=4, value=img.get('format', ''))
        c4.font = Font(name='Calibri', size=10, color='444444')
        c4.fill = rf; c4.alignment = Alignment(horizontal='center', vertical='center'); c4.border = bd
        ws.row_dimensions[r].height = 20
    output = io.BytesIO()
    wb.save(output); output.seek(0)
    fname = f'amazon_image_urls_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=fname)

@app.route('/health')
def health():
    return jsonify({'status': 'ok'})

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=os.environ.get('FLASK_ENV') == 'development')
